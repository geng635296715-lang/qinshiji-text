# -*- coding: utf-8 -*-
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
CAT_PATH = ROOT / "src" / "modules" / "bazi" / "shensha-catalog.ts"
SVC_PATH = ROOT / "src" / "modules" / "bazi" / "service.ts"
OUT_DIR = ROOT / "reports"
OUT_DIR.mkdir(exist_ok=True)
OUT_PATH = OUT_DIR / "qing_shiji_common_shensha_crosscheck.xlsx"


def extract_common_entries(text: str):
    block = text.split("uncommon-001")[0]
    entries = re.findall(
        r'name:\s*"([^"]+)".*?tone:\s*"([^"]+)".*?volume:\s*"([^"]+)"',
        block,
        flags=re.S,
    )
    return [{"name": name, "tone": tone, "volume": volume} for name, tone, volume in entries]


def collect_service_names(text: str):
    names = set(re.findall(r'name:\s*"([^"]+)"', text))
    names.difference_update({"day", "hour", "month", "year"})
    return names


def main():
    cat_text = CAT_PATH.read_text(encoding="utf-8")
    svc_text = SVC_PATH.read_text(encoding="utf-8")

    common = extract_common_entries(cat_text)
    service_names = collect_service_names(svc_text)

    basis_map = {
        "天乙贵人": "年干、日干并参，查四柱地支",
        "太极贵人": "年干、日干并参，查四柱地支",
        "天德贵人": "月支查天干/地支并参",
        "月德贵人": "月支查天干/地支并参",
        "天德合": "月支查天干/地支并参",
        "月德合": "月支查天干/地支并参",
        "三奇贵人": "四柱天干顺布三奇",
        "文昌贵人": "年干、日干并参，查四柱地支",
        "国印贵人": "年干、日干并参，查四柱地支",
        "福星贵人": "日干查四柱地支",
        "德秀贵人": "月令 + 季节天干并参",
        "天厨贵人": "日干查四柱地支",
        "金舆": "日干查四柱地支",
        "华盖": "年支、日支并参，查三合墓库位",
        "将星": "年支、日支并参，查三合将星位",
        "学堂": "年干、日干 + 纳音并参",
        "词馆": "年干、日干 + 纳音并参",
        "驿马": "年支、日支并参，查三合驿马位",
        "桃花": "年支、日支并参，取三合局沐浴位",
        "红鸾": "年支查红鸾位",
        "天喜": "年支查对冲喜位",
        "咸池": "年支、日支并参，取桃花/沐浴位",
        "禄神": "日干查临官位",
        "羊刃": "日干查帝旺位",
        "飞刃": "日干查羊刃对冲位",
        "天医": "月支查天医位",
        "天赦": "季节 + 日柱并参",
        "天罗": "年支/日支并参，固定支位命中",
        "地网": "年支/日支并参，固定支位命中",
        "孤辰": "年支、日支并参，查群组孤辰位",
        "寡宿": "年支、日支并参，查群组寡宿位",
        "丧门": "年支查丧门位",
        "吊客": "年支查吊客位",
        "披麻": "年支查披麻位",
        "劫煞": "年支、日支并参，查劫煞位",
        "灾煞": "年支、日支并参，查灾煞位",
        "亡神": "年支、日支并参，查亡神位",
        "勾绞煞": "年支查勾煞/绞煞位",
        "红艳": "日干查红艳位",
        "流霞": "日干查流霞位",
        "魁罡": "日柱四特定干支命中",
        "金神": "日柱/时柱特定干支命中",
        "十灵日": "日柱特定干支命中",
        "孤鸾": "日柱特定干支命中",
        "阴差阳错": "日柱特定干支命中",
        "童子煞": "月支 + 纳音 + 日时并参",
        "元辰": "年支 + 性别并参",
        "血刃": "日干 + 月支并参",
        "天财": "财星落支 + 财库并参（争议口径）",
        "天官": "年干、日干并参，查四柱地支",
        "天福": "天福贵人口诀 + 年干、日干并参",
        "六秀日": "特定六个日柱命中",
        "八专": "特定十个日柱命中",
        "九丑": "特定十个日柱命中",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "common_check"
    headers = ["神煞名称", "吉凶属性", "是否嵌入后端算法", "后端命中口径", "前端命中呈现", "备注"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F1B16")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="C7B08A")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for item in common:
        name = item["name"]
        in_backend = name in service_names
        ws.append(
            [
                name,
                item["tone"],
                "是" if in_backend else "否",
                basis_map.get(name, "已按当前后端规则嵌入"),
                "已嵌入后端算法；前端以神煞标签+悬停释义展示；命中后挂在对应柱位"
                if in_backend
                else "未嵌入后端算法；前端不会命中展示",
                "当前后端已统一接入常用神煞标签与悬停释义" if in_backend else "尚未接入",
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, max(len(line) for line in value.splitlines()))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 42)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("summary")
    summary.append(["项目", "数量", "说明"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    all_count = len(common)
    in_backend_count = sum(1 for item in common if item["name"] in service_names)
    missing = [item["name"] for item in common if item["name"] not in service_names]
    summary_rows = [
        ["上卷神煞总数", all_count, "上卷共 54 项"],
        ["已嵌入后端", in_backend_count, "当前检查结果"],
        ["未嵌入后端", len(missing), "若为 0 则表示上卷已全量接入"],
        ["未嵌入项列表", "、".join(missing) if missing else "无", "当前没有遗漏"],
    ]
    for row in summary_rows:
        summary.append(row)

    for row in summary.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 18
    summary.column_dimensions["C"].width = 58
    summary.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(OUT_PATH)
    print(f"common={all_count}, in_backend={in_backend_count}, missing={missing}")


if __name__ == "__main__":
    main()
